import csv
import re
import shutil
from datetime import datetime
from ipaddress import IPv4Address
from pathlib import Path
from typing import List, Optional, Tuple

import tkinter as tk
from tkinter import filedialog, messagebox, ttk

try:
    import openpyxl  # for xlsx
except Exception:
    openpyxl = None


# IPv4 matcher
IP_RE = re.compile(
    r"\b(?:(?:25[0-5]|2[0-4]\d|1?\d?\d)\.){3}(?:25[0-5]|2[0-4]\d|1?\d?\d)\b"
)


def extract_ips_from_text(text: str) -> List[str]:
    return IP_RE.findall(text)


def normalize_unique_ips(ips: List[str]) -> List[str]:
    """Validate IPv4, normalize, unique-preserve order."""
    seen = set()
    out: List[str] = []
    for s in ips:
        try:
            ip = str(IPv4Address(s))
        except Exception:
            continue
        if ip not in seen:
            seen.add(ip)
            out.append(ip)
    return out


def read_ips_from_txt(path: Path) -> List[str]:
    text = path.read_text(encoding="utf-8", errors="ignore")
    return extract_ips_from_text(text)


def read_ips_from_csv(path: Path) -> List[str]:
    ips: List[str] = []
    with path.open("r", encoding="utf-8", errors="ignore", newline="") as f:
        sample = f.read(4096)
        f.seek(0)
        try:
            dialect = csv.Sniffer().sniff(sample)
        except Exception:
            dialect = csv.excel

        reader = csv.reader(f, dialect)
        rows = list(reader)
        if not rows:
            return []

        # Detect header and "ip" column if present
        header = [c.strip().lower() for c in rows[0]]
        ip_col = None
        for i, name in enumerate(header):
            if name in ("ip", "ipaddress", "ip_address", "address", "host"):
                ip_col = i
                break

        looks_like_header = any(any(ch.isalpha() for ch in cell) for cell in rows[0])
        start_idx = 1 if (looks_like_header and ip_col is not None) else 0

        for r in rows[start_idx:]:
            if not r:
                continue
            if ip_col is not None and ip_col < len(r):
                ips.extend(extract_ips_from_text(str(r[ip_col])))
            else:
                ips.extend(extract_ips_from_text(" ".join(map(str, r))))
    return ips


def read_ips_from_xlsx(path: Path) -> List[str]:
    if openpyxl is None:
        raise RuntimeError("openpyxl не установлен (нужен для .xlsx). Установи: pip install openpyxl")

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active

    first_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
    ip_col = None
    if first_row:
        header = [str(c).strip().lower() if c is not None else "" for c in first_row]
        for i, name in enumerate(header):
            if name in ("ip", "ipaddress", "ip_address", "address", "host"):
                ip_col = i
                break

    ips: List[str] = []
    for row in ws.iter_rows(values_only=True):
        if not row:
            continue
        # Skip header row if we detected an ip column
        if first_row == row and ip_col is not None:
            continue

        if ip_col is not None and ip_col < len(row):
            cell = row[ip_col]
            if cell is not None:
                ips.extend(extract_ips_from_text(str(cell)))
        else:
            joined = " ".join("" if c is None else str(c) for c in row)
            ips.extend(extract_ips_from_text(joined))

    wb.close()
    return ips


def parse_iprangegroups_line(ini_text: str) -> Tuple[Optional[str], bool]:
    """
    Returns (current_value_without_quotes, quoted)
    """
    m = re.search(r"(?im)^\s*iprangegroups\s*=\s*(.*)$", ini_text)
    if not m:
        return None, False

    raw = m.group(1).strip()
    if len(raw) >= 2 and raw[0] == '"' and raw[-1] == '"':
        return raw[1:-1], True
    return raw, False


def chunk_list(items: List[str], chunk_size: int) -> List[List[str]]:
    if chunk_size <= 0:
        return [items]
    return [items[i : i + chunk_size] for i in range(0, len(items), chunk_size)]


def build_iprangegroups_value_ips_only(
    ips: List[str],
    quoted: bool,
    group_prefix: str = "#LAN",
    group_sep: str = ";",
    kv_sep: str = ":",
    ip_sep: str = ",",
    chunk_size: int = 0,
) -> str:
    """
    Produce BTC Tools style:
      ipRangeGroups="#LAN:ip1,ip2,ip3"
    If chunk_size > 0 and ips are many:
      ipRangeGroups="#LAN1:...;#LAN2:...;#LAN3:..."
    """
    if not ips:
        value = f"{group_prefix}{kv_sep}"
        return f"\"{value}\"" if quoted else value

    chunks = chunk_list(ips, chunk_size) if chunk_size else [ips]

    parts: List[str] = []
    for idx, ch in enumerate(chunks, start=1):
        name = group_prefix if len(chunks) == 1 else f"{group_prefix}{idx}"
        parts.append(f"{name}{kv_sep}{ip_sep.join(ch)}")

    value = group_sep.join(parts)
    return f"\"{value}\"" if quoted else value


def patch_ini(ini_path: Path, new_value: str) -> Tuple[bool, str]:
    """
    Patch/insert ipRangeGroups=... preserving line endings.
    """
    lines = ini_path.read_text(encoding="utf-8", errors="ignore").splitlines(True)
    out_lines: List[str] = []
    found = False
    changed = False

    for line in lines:
        if line.strip().lower().startswith("iprangegroups="):
            found = True
            old = line.split("=", 1)[1].rstrip("\r\n")
            if old != new_value:
                newline = "\r\n" if line.endswith("\r\n") else "\n"
                out_lines.append("ipRangeGroups=" + new_value + newline)
                changed = True
            else:
                out_lines.append(line)
            continue
        out_lines.append(line)

    if not found:
        # Append at end
        if out_lines and not out_lines[-1].endswith(("\n", "\r\n")):
            out_lines.append("\n")
        out_lines.append("ipRangeGroups=" + new_value + "\n")
        changed = True

    if changed:
        ini_path.write_text("".join(out_lines), encoding="utf-8", errors="ignore")
        return True, "Готово: ipRangeGroups обновлён."
    return False, "Ничего не изменилось: ipRangeGroups уже совпадает."


class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("BTC Tools ini patcher (ipRangeGroups → IP list)")
        self.geometry("820x520")
        self.minsize(820, 520)

        self.ini_path_var = tk.StringVar()
        self.ip_path_var = tk.StringVar()

        # If many IPs, split to groups (#LAN1, #LAN2...) by this size.
        # 0 = no split
        self.chunk_size_var = tk.StringVar(value="1000")

        self.preview = tk.Text(self, height=12, wrap="word")
        self.preview.configure(state="disabled")

        self._build_ui()

    def _build_ui(self):
        pad = {"padx": 10, "pady": 6}
        frm = ttk.Frame(self)
        frm.pack(fill="both", expand=True)

        # INI
        r1 = ttk.Frame(frm)
        r1.pack(fill="x", **pad)
        ttk.Label(r1, text="BTC_Tools.ini:").pack(side="left")
        ttk.Entry(r1, textvariable=self.ini_path_var).pack(side="left", fill="x", expand=True, padx=8)
        ttk.Button(r1, text="Выбрать...", command=self.pick_ini).pack(side="left")

        # IP file
        r2 = ttk.Frame(frm)
        r2.pack(fill="x", **pad)
        ttk.Label(r2, text="Файл с IP (txt/csv/xlsx):").pack(side="left")
        ttk.Entry(r2, textvariable=self.ip_path_var).pack(side="left", fill="x", expand=True, padx=8)
        ttk.Button(r2, text="Выбрать...", command=self.pick_ipfile).pack(side="left")

        # Chunking
        r3 = ttk.Frame(frm)
        r3.pack(fill="x", **pad)
        ttk.Label(r3, text="Разбивка по группам (IP в одной группе):").pack(side="left")
        e = ttk.Entry(r3, textvariable=self.chunk_size_var, width=10)
        e.pack(side="left", padx=8)
        ttk.Label(r3, text="(0 = не дробить; напр. 500/1000, если IP очень много)").pack(side="left")

        # Buttons
        r4 = ttk.Frame(frm)
        r4.pack(fill="x", **pad)
        ttk.Button(r4, text="Сформировать превью", command=self.refresh_preview).pack(side="left")
        ttk.Button(r4, text="Патчить ini (с бэкапом)", command=self.do_patch).pack(side="left", padx=10)

        ttk.Label(frm, text="Превью нового значения ipRangeGroups:").pack(anchor="w", padx=12, pady=(12, 0))
        self.preview.pack(fill="both", expand=True, padx=12, pady=(6, 12))

        ttk.Label(
            frm,
            text="Формат: ipRangeGroups=\"#LAN:ip1,ip2,...\" (или #LAN1/#LAN2 при дроблении).",
        ).pack(anchor="w", padx=12, pady=(0, 8))

    def pick_ini(self):
        p = filedialog.askopenfilename(
            title="Выбери BTC_Tools.ini",
            filetypes=[("INI files", "*.ini"), ("All files", "*.*")]
        )
        if p:
            self.ini_path_var.set(p)
            self.refresh_preview()

    def pick_ipfile(self):
        p = filedialog.askopenfilename(
            title="Выбери файл с IP",
            filetypes=[("Supported", "*.txt *.csv *.xlsx"), ("Text", "*.txt"), ("CSV", "*.csv"), ("Excel", "*.xlsx"), ("All files", "*.*")]
        )
        if p:
            self.ip_path_var.set(p)
            self.refresh_preview()

    def _load_ips(self) -> List[str]:
        ip_path = Path(self.ip_path_var.get().strip())
        if not ip_path.exists():
            raise FileNotFoundError("Файл с IP не найден.")

        ext = ip_path.suffix.lower()
        if ext == ".txt":
            ips = read_ips_from_txt(ip_path)
        elif ext == ".csv":
            ips = read_ips_from_csv(ip_path)
        elif ext == ".xlsx":
            ips = read_ips_from_xlsx(ip_path)
        else:
            ips = read_ips_from_txt(ip_path)

        ips = normalize_unique_ips(ips)
        if not ips:
            raise RuntimeError("IP не найдены в файле.")
        return ips

    def _get_chunk_size(self) -> int:
        raw = self.chunk_size_var.get().strip()
        if not raw:
            return 0
        try:
            v = int(raw)
            if v < 0:
                v = 0
            return v
        except Exception:
            raise RuntimeError("Разбивка должна быть числом (например 0, 500, 1000).")

    def _make_new_value(self) -> str:
        ini_path = Path(self.ini_path_var.get().strip())
        if not ini_path.exists():
            raise FileNotFoundError("BTC_Tools.ini не найден.")

        ini_text = ini_path.read_text(encoding="utf-8", errors="ignore")
        _, quoted = parse_iprangegroups_line(ini_text)

        # Если строки нет — по умолчанию делаем в кавычках (как в твоём ini)
        if _ is None:
            quoted = True

        ips = self._load_ips()
        chunk_size = self._get_chunk_size()

        # BTC Tools стиль: #LAN:ip,ip,ip  (или #LAN1/#LAN2 если дробим)
        return build_iprangegroups_value_ips_only(
            ips=ips,
            quoted=quoted,
            group_prefix="#LAN",
            group_sep=";",
            kv_sep=":",
            ip_sep=",",
            chunk_size=chunk_size,
        )

    def refresh_preview(self):
        try:
            if not self.ini_path_var.get().strip() or not self.ip_path_var.get().strip():
                return
            new_value = self._make_new_value()
            self.preview.configure(state="normal")
            self.preview.delete("1.0", "end")
            self.preview.insert("1.0", "ipRangeGroups=" + new_value)
            self.preview.configure(state="disabled")
        except Exception as e:
            self.preview.configure(state="normal")
            self.preview.delete("1.0", "end")
            self.preview.insert("1.0", f"[Ошибка превью] {e}")
            self.preview.configure(state="disabled")

    def do_patch(self):
        try:
            ini_path = Path(self.ini_path_var.get().strip())
            if not ini_path.exists():
                raise FileNotFoundError("BTC_Tools.ini не найден.")

            new_value = self._make_new_value()

            # backup
            ts = datetime.now().strftime("%Y%m%d_%H%M%S")
            backup_path = ini_path.with_suffix(ini_path.suffix + f".bak_{ts}")
            shutil.copy2(ini_path, backup_path)

            changed, msg = patch_ini(ini_path, new_value)
            self.refresh_preview()

            if changed:
                messagebox.showinfo("Успех", f"{msg}\n\nБэкап: {backup_path.name}")
            else:
                messagebox.showinfo("Ок", f"{msg}\n\nБэкап всё равно создан: {backup_path.name}")

        except Exception as e:
            messagebox.showerror("Ошибка", str(e))


def main():
    app = App()
    app.mainloop()


if __name__ == "__main__":
    main()
